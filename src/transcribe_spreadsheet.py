from fastapi import FastAPI, UploadFile, File, Form
import io
import logging
from typing import Optional, List, Dict, Any
import openpyxl
import csv
import json
import pandas as pd
import requests
from datetime import datetime

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

@app.get("/health")
def health():
    return {"ok": True, "service": "spreadsheet-processor"}

@app.get("/test")
def test():
    """Endpoint de teste para verificar funcionamento"""
    return {
        "status": "ok",
        "message": "Serviço de processamento de planilhas funcionando",
        "version": "1.0.0",
        "supported_formats": ["xlsx", "xls", "csv"]
    }

@app.post("/debug-file")
async def debug_file(file: UploadFile = File(...)):
    """Endpoint para debug de arquivos - mostra informações detalhadas"""
    try:
        data = await file.read()
        
        # Informações básicas
        info = {
            "filename": file.filename,
            "content_type": file.content_type,
            "size_bytes": len(data),
            "first_100_bytes": data[:100].hex() if len(data) >= 100 else data.hex(),
            "first_100_chars": data[:100].decode('utf-8', errors='ignore') if len(data) >= 100 else data.decode('utf-8', errors='ignore'),
        }
        
        # Detectar tipo usando a nova função
        detected_type = detect_file_type(file.filename, file.content_type, data)
        info["detected_type"] = detected_type
        
        # Tentar processar para ver se funciona
        try:
            if detected_type == "CSV":
                sheets_data = process_csv(data)
                info["csv_processing"] = "success"
                info["csv_sheets"] = list(sheets_data.keys())
            else:
                sheets_data = process_excel(data)
                info["excel_processing"] = "success"
                info["excel_sheets"] = list(sheets_data.keys())
        except Exception as e:
            info["processing_error"] = str(e)
        
        return {
            "status": "success",
            "file_info": info
        }
        
    except Exception as e:
        return {
            "status": "error",
            "error": str(e)
        }

def detect_file_type(filename: str, content_type: str = None, data: bytes = None) -> str:
    """Detecta o tipo de planilha baseado na extensão, MIME type e conteúdo"""
    
    # 1. Tentar detectar pelo MIME type primeiro
    if content_type:
        if 'spreadsheet' in content_type or 'excel' in content_type:
            if 'xlsx' in content_type or 'openxml' in content_type:
                return "XLSX"
            elif 'xls' in content_type:
                return "XLS"
        elif 'csv' in content_type or 'text/csv' in content_type:
            return "CSV"
    
    # 2. Tentar detectar pelos primeiros bytes do arquivo
    if data:
        # Verificar se é CSV (texto com vírgulas, ponto e vírgula, ou tabs)
        try:
            text_sample = data[:1000].decode('utf-8', errors='ignore')
            # Verificar padrões típicos de CSV
            if ',' in text_sample or ';' in text_sample or '\t' in text_sample:
                # Verificar se tem quebras de linha (indicativo de CSV)
                if '\n' in text_sample or '\r' in text_sample:
                    return "CSV"
        except:
            pass
        
        # Verificar se é Excel (XLSX é um ZIP)
        if data.startswith(b'PK'):
            return "XLSX"
        # Verificar se é Excel antigo (XLS)
        elif data.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'):
            return "XLS"
    
    # 3. Detectar pela extensão do arquivo
    filename_lower = filename.lower()
    if filename_lower.endswith('.xlsx') or filename_lower.endswith('.xlsm'):
        return "XLSX"
    elif filename_lower.endswith('.xls'):
        return "XLS"
    elif filename_lower.endswith('.csv'):
        return "CSV"
    else:
        return "UNKNOWN"

def process_csv(data: bytes) -> Dict[str, Any]:
    """Processa arquivo CSV"""
    try:
        # Tentar detectar o encoding
        text = data.decode('utf-8')
    except UnicodeDecodeError:
        try:
            text = data.decode('latin-1')
        except:
            text = data.decode('utf-8', errors='ignore')
    
    # Ler CSV
    lines = text.split('\n')
    reader = csv.DictReader(lines)
    
    rows = []
    for row in reader:
        rows.append(row)
    
    return {
        "Planilha": {
            "headers": list(rows[0].keys()) if rows else [],
            "rows": rows,
            "total_rows": len(rows)
        }
    }

def process_excel(data: bytes) -> Dict[str, Any]:
    """Processa arquivo Excel (xlsx/xls) com múltiplas abas"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        logger.info(f"Excel carregado com sucesso. Abas: {workbook.sheetnames}")
    except Exception as e:
        logger.error(f"Erro ao carregar Excel: {e}")
        raise
    
    all_sheets_data = {}
    
    for sheet_name in workbook.sheetnames:
        logger.info(f"Processando aba: {sheet_name}")
        sheet = workbook[sheet_name]
        
        # Extrair dados da aba
        rows_data = []
        headers = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            # Pular linhas completamente vazias
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            if idx == 1:
                # Primeira linha como cabeçalho
                headers = [str(cell) if cell is not None else f"Coluna_{i}" for i, cell in enumerate(row, 1)]
            else:
                # Criar dicionário para cada linha
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Converter valores para strings legíveis
                        if cell is None:
                            value = ""
                        elif isinstance(cell, (int, float)):
                            value = str(cell)
                        elif isinstance(cell, datetime):
                            value = cell.strftime("%Y-%m-%d %H:%M:%S")
                        else:
                            value = str(cell)
                        
                        row_dict[headers[i]] = value
                
                rows_data.append(row_dict)
        
        all_sheets_data[sheet_name] = {
            "headers": headers,
            "rows": rows_data,
            "total_rows": len(rows_data)
        }
        
        logger.info(f"Aba '{sheet_name}' processada: {len(rows_data)} linhas")
    
    workbook.close()
    return all_sheets_data

def format_for_rag(sheets_data: Dict[str, Any], include_metadata: bool = True) -> str:
    """
    Formata os dados da planilha em texto estruturado para RAG.
    Cada aba se torna uma seção com seus dados organizados.
    """
    rag_text = []
    
    if include_metadata:
        rag_text.append(f"# DADOS DA PLANILHA\n")
        rag_text.append(f"Data de processamento: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        rag_text.append(f"Total de abas: {len(sheets_data)}\n")
        rag_text.append("-" * 80 + "\n\n")
    
    for sheet_name, sheet_data in sheets_data.items():
        rag_text.append(f"## ABA: {sheet_name}\n")
        rag_text.append(f"Total de registros: {sheet_data['total_rows']}\n")
        rag_text.append(f"Colunas: {', '.join(sheet_data['headers'])}\n\n")
        
        # Formatar cada linha de dados
        for idx, row in enumerate(sheet_data['rows'], 1):
            rag_text.append(f"### Registro {idx}\n")
            for key, value in row.items():
                if value and str(value).strip():  # Só incluir valores não vazios
                    rag_text.append(f"- **{key}**: {value}\n")
            rag_text.append("\n")
        
        rag_text.append("-" * 80 + "\n\n")
    
    return ''.join(rag_text)

def format_for_rag_compact(sheets_data: Dict[str, Any]) -> str:
    """
    Formata os dados de forma mais compacta, ideal para RAG com muitos dados.
    Usa formato de tabela markdown para economia de tokens.
    """
    rag_text = []
    
    rag_text.append(f"# DADOS DA PLANILHA\n\n")
    
    for sheet_name, sheet_data in sheets_data.items():
        rag_text.append(f"## {sheet_name}\n")
        rag_text.append(f"*{sheet_data['total_rows']} registros*\n\n")
        
        if not sheet_data['rows']:
            rag_text.append("*(Sem dados)*\n\n")
            continue
        
        # Criar tabela markdown
        headers = sheet_data['headers']
        rag_text.append("| " + " | ".join(headers) + " |\n")
        rag_text.append("|" + "|".join(["---" for _ in headers]) + "|\n")
        
        for row in sheet_data['rows']:
            values = [str(row.get(h, "")) for h in headers]
            rag_text.append("| " + " | ".join(values) + " |\n")
        
        rag_text.append("\n")
    
    return ''.join(rag_text)

def format_as_json(sheets_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Formata os dados como JSON estruturado para envio direto ao n8n.
    """
    return {
        "processed_at": datetime.now().isoformat(),
        "total_sheets": len(sheets_data),
        "sheets": sheets_data
    }

@app.post("/extract-spreadsheet")
async def extract_spreadsheet(
    file: UploadFile = File(...),
    format: str = Form("markdown"),  # "markdown", "markdown-compact", "json"
    include_metadata: bool = Form(True)
):
    """
    Extrai e processa dados de planilha com múltiplas abas.
    
    Parâmetros:
    - file: Arquivo da planilha (xlsx, xls, csv)
    - format: Formato de saída ("markdown", "markdown-compact", "json")
    - include_metadata: Incluir metadados no output (apenas para markdown)
    
    Retorna dados estruturados prontos para envio ao RAG.
    """
    try:
        logger.info(f"Iniciando processamento da planilha: {file.filename}")
        
        # Ler dados do arquivo
        data = await file.read()
        logger.info(f"Arquivo lido com sucesso. Tamanho: {len(data)} bytes")
        
        if len(data) == 0:
            return {
                "status": "error",
                "error": "Arquivo vazio"
            }
        
        # Detectar tipo de arquivo
        file_type = detect_file_type(
            file.filename, 
            file.content_type, 
            data
        )
        logger.info(f"Tipo de arquivo detectado: {file_type} (filename: {file.filename}, content_type: {file.content_type})")
        
        if file_type == "UNKNOWN":
            return {
                "status": "error",
                "error": f"Formato de arquivo não suportado. Use xlsx, xls ou csv. Detectado: filename='{file.filename}', content_type='{file.content_type}'"
            }
        
        # Processar planilha
        if file_type == "CSV":
            sheets_data = process_csv(data)
        else:  # XLSX ou XLS
            sheets_data = process_excel(data)
        
        # Formatar output baseado no parâmetro format
        if format == "json":
            output = format_as_json(sheets_data)
            return {
                "status": "success",
                "file_type": file_type,
                "data": output
            }
        elif format == "markdown-compact":
            text = format_for_rag_compact(sheets_data)
        else:  # markdown (padrão)
            text = format_for_rag(sheets_data, include_metadata)
        
        return {
            "status": "success",
            "file_type": file_type,
            "text": text,
            "sheets": list(sheets_data.keys()),
            "total_sheets": len(sheets_data),
            "total_rows": sum(s['total_rows'] for s in sheets_data.values())
        }
        
    except Exception as e:
        logger.error(f"Erro ao processar planilha: {e}", exc_info=True)
        return {
            "status": "error",
            "error": f"Erro interno: {str(e)}"
        }

@app.post("/extract-for-n8n")
async def extract_for_n8n(file: UploadFile = File(...)):
    """
    Extrai dados da planilha otimizado para processamento no n8n.
    
    Retorna um array onde cada objeto é uma aba da planilha com:
    - Metadados da aba (nome, total de linhas, colunas)
    - Dados completos da aba
    
    Ideal para processar no n8n com loop e vetorização posterior.
    
    Parâmetros:
    - file: Arquivo da planilha (xlsx, xls, csv)
    
    Retorna: Array de objetos, um para cada aba
    """
    try:
        logger.info(f"Processando planilha para n8n: {file.filename}")
        
        # Ler dados do arquivo
        data = await file.read()
        logger.info(f"Arquivo lido: {len(data)} bytes")
        
        if len(data) == 0:
            return {
                "status": "error",
                "error": "Arquivo vazio"
            }
        
        # Detectar tipo de arquivo
        file_type = detect_file_type(
            file.filename, 
            file.content_type, 
            data
        )
        logger.info(f"Tipo detectado: {file_type} (filename: {file.filename}, content_type: {file.content_type})")
        
        if file_type == "UNKNOWN":
            return {
                "status": "error",
                "error": f"Formato não suportado. Use xlsx, xls ou csv. Detectado: filename='{file.filename}', content_type='{file.content_type}'"
            }
        
        # Processar planilha
        if file_type == "CSV":
            sheets_data = process_csv(data)
        else:
            sheets_data = process_excel(data)
        
        # Montar array de abas para n8n
        sheets_array = []
        
        for sheet_name, sheet_data in sheets_data.items():
            sheet_obj = {
                "sheet_name": sheet_name,
                "metadata": {
                    "total_rows": sheet_data['total_rows'],
                    "total_columns": len(sheet_data['headers']),
                    "columns": sheet_data['headers'],
                    "source_file": file.filename,
                    "file_type": file_type,
                    "processed_at": datetime.now().isoformat()
                },
                "data": sheet_data['rows']
            }
            sheets_array.append(sheet_obj)
        
        logger.info(f"Processamento concluído: {len(sheets_array)} abas")
        
        return {
            "status": "success",
            "filename": file.filename,
            "file_type": file_type,
            "total_sheets": len(sheets_array),
            "total_rows": sum(s['metadata']['total_rows'] for s in sheets_array),
            "sheets": sheets_array
        }
        
    except Exception as e:
        logger.error(f"Erro ao processar: {e}", exc_info=True)
        return {
            "status": "error",
            "error": f"Erro interno: {str(e)}"
        }

@app.post("/extract-and-send-to-n8n")
async def extract_and_send_to_n8n(
    file: UploadFile = File(...),
    n8n_webhook_url: str = Form(...)
):
    """
    Extrai dados da planilha e envia diretamente para o webhook do n8n.
    
    Envia um array onde cada objeto é uma aba da planilha completa.
    
    Parâmetros:
    - file: Arquivo da planilha
    - n8n_webhook_url: URL do webhook do n8n
    
    Retorna confirmação do envio.
    """
    try:
        logger.info(f"Processando e enviando para n8n: {file.filename}")
        
        # Usar o endpoint otimizado
        data = await file.read()
        file_type = detect_file_type(
            file.filename, 
            file.content_type, 
            data
        )
        logger.info(f"Tipo detectado: {file_type} (filename: {file.filename}, content_type: {file.content_type})")
        
        if file_type == "CSV":
            sheets_data = process_csv(data)
        else:
            sheets_data = process_excel(data)
        
        # Montar array de abas
        sheets_array = []
        for sheet_name, sheet_data in sheets_data.items():
            sheet_obj = {
                "sheet_name": sheet_name,
                "metadata": {
                    "total_rows": sheet_data['total_rows'],
                    "total_columns": len(sheet_data['headers']),
                    "columns": sheet_data['headers'],
                    "source_file": file.filename,
                    "file_type": file_type,
                    "processed_at": datetime.now().isoformat()
                },
                "data": sheet_data['rows']
            }
            sheets_array.append(sheet_obj)
        
        # Preparar payload
        payload = {
            "status": "success",
            "filename": file.filename,
            "file_type": file_type,
            "total_sheets": len(sheets_array),
            "total_rows": sum(s['metadata']['total_rows'] for s in sheets_array),
            "sheets": sheets_array
        }
        
        # Enviar para n8n
        logger.info(f"Enviando para n8n: {n8n_webhook_url}")
        response = requests.post(
            n8n_webhook_url,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=60
        )
        
        response.raise_for_status()
        
        return {
            "status": "success",
            "message": "Dados enviados para n8n com sucesso",
            "n8n_response_status": response.status_code,
            "sheets_processed": [s['sheet_name'] for s in sheets_array],
            "total_rows": sum(s['metadata']['total_rows'] for s in sheets_array)
        }
        
    except requests.exceptions.RequestException as e:
        logger.error(f"Erro ao enviar para n8n: {e}")
        return {
            "status": "error",
            "error": f"Erro ao enviar para n8n: {str(e)}"
        }
    except Exception as e:
        logger.error(f"Erro geral: {e}", exc_info=True)
        return {
            "status": "error",
            "error": f"Erro interno: {str(e)}"
        }

@app.post("/analyze-spreadsheet")
async def analyze_spreadsheet(file: UploadFile = File(...)):
    """
    Analisa a estrutura da planilha sem processar todos os dados.
    Útil para preview e validação.
    """
    try:
        data = await file.read()
        file_type = detect_file_type(
            file.filename, 
            file.content_type, 
            data
        )
        logger.info(f"Tipo detectado: {file_type} (filename: {file.filename}, content_type: {file.content_type})")
        
        if file_type == "CSV":
            sheets_data = process_csv(data)
        else:
            sheets_data = process_excel(data)
        
        # Criar análise resumida
        analysis = {
            "filename": file.filename,
            "file_type": file_type,
            "total_sheets": len(sheets_data),
            "sheets_info": []
        }
        
        for sheet_name, sheet_data in sheets_data.items():
            sheet_info = {
                "name": sheet_name,
                "total_rows": sheet_data['total_rows'],
                "total_columns": len(sheet_data['headers']),
                "columns": sheet_data['headers'],
                "sample_rows": sheet_data['rows'][:3]  # Primeiras 3 linhas como amostra
            }
            analysis["sheets_info"].append(sheet_info)
        
        return {
            "status": "success",
            "analysis": analysis
        }
        
    except Exception as e:
        logger.error(f"Erro ao analisar planilha: {e}")
        return {
            "status": "error",
            "error": str(e)
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
